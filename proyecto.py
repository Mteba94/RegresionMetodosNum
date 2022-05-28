import pandas as pd
import requests
import os
from tabula import read_pdf
import tabula
import matplotlib.pyplot as plt
import openpyxl
from cProfile import label
from statistics import linear_regression
from sklearn import linear_model
from sklearn.model_selection import train_test_split

#Eliminar los archivos para volver a descargarlos actualizados
os.remove('Informe-Mensual-ABRIL-2022-SC.pdf')
os.remove('Informe.csv')
os.remove('data.xlsx')
os.remove('SuperiorA.xlsx')
os.remove('Superior.xlsx')

#Descargar informacion de la pagina en formato PDF

url = "https://mem.gob.gt/wp-content/uploads/2022/05/Informe-Mensual-ABRIL-2022-SC.pdf"

resp = requests.get(url)

output = open('Informe-Mensual-ABRIL-2022-SC.pdf', 'wb')
output.write(resp.content)
output.close()

#Convertir PDF a CSV

filename = 'Informe-Mensual-ABRIL-2022-SC.pdf'
df = read_pdf( input_path=filename, pages='1')
df = read_pdf( input_path=filename, pages='2')

tabula.convert_into("Informe-Mensual-ABRIL-2022-SC.pdf", "Informe.csv", output_format="csv", pages='1')
#tabula.convert_into("Informe-Mensual-ABRIL-2022-SC.pdf", "Informe2.csv", output_format="csv", pages='2')

#Convertir CSV a Excel

df = pd.read_csv('Informe.csv', encoding='ISO-8859-1')
df.to_excel('data.xlsx', index=False)

wb = openpyxl.load_workbook('data.xlsx')

sheet = wb['Sheet1']

sheet.delete_rows(15,15)

wb.save('SuperiorA.xlsx')

wb = openpyxl.load_workbook('data.xlsx')

sheet = wb['Sheet1']

sheet.delete_rows(1,15)

wb.save('Superior.xlsx')

#Convierto el segundo csv
#df = pd.read_csv('Informe2.csv', encoding='ISO-8859-1')
#df.to_excel('data2.xlsx', index=False)

#wb = openpyxl.load_workbook('data2.xlsx')

#sheet = wb['Sheet1']

#sheet.delete_rows(1,1)
#sheet.delete_rows(15,15)

#wb.save('Regular.xlsx')

#Realiza graficas para apreciar la informacion

readworkbook = pd.read_excel('SuperiorA.xlsx')
readworkbook.head()

fuente1 = {'family':'monospace','style':'italic','weight':'bold','color':'red','size':17}

fig = plt.figure(figsize=(14,14))
plt.scatter(readworkbook['MES/AÑO'],readworkbook['2018'])
plt.scatter(readworkbook['MES/AÑO'],readworkbook['2019'])
plt.scatter(readworkbook['MES/AÑO'],readworkbook['2020'])
plt.scatter(readworkbook['MES/AÑO'],readworkbook['2021'])
plt.scatter(readworkbook['MES/AÑO'],readworkbook['2022'])
#plt.scatter(readworkbook['MES/AÑO'],readworkbook['2006'])
#plt.scatter(readworkbook['MES/AÑO'],readworkbook['2007'])
#plt.scatter(readworkbook['MES/AÑO'],readworkbook['2008'])
plt.plot(readworkbook['MES/AÑO'],readworkbook['2018'], label="2018")
plt.plot(readworkbook['MES/AÑO'],readworkbook['2019'], label="2019")
plt.plot(readworkbook['MES/AÑO'],readworkbook['2020'], label="2020")
plt.plot(readworkbook['MES/AÑO'],readworkbook['2021'], label="2021")
plt.plot(readworkbook['MES/AÑO'],readworkbook['2022'], label="2022")
#plt.plot(readworkbook['MES/AÑO'],readworkbook['2006'], label="2006")
#plt.plot(readworkbook['MES/AÑO'],readworkbook['2007'], label="2007")
#plt.plot(readworkbook['MES/AÑO'],readworkbook['2008'], label="2008")
plt.legend(loc="upper left")
plt.xlabel('MES', fontdict=fuente1) 
plt.ylabel('PRECIO Q', fontdict=fuente1)
plt.title("Gasolina Superior con Aditivo")
plt.grid
plt.show()

readworkbook = pd.read_excel('Superior.xlsx')
readworkbook.head()

fuente1 = {'family':'monospace','style':'italic','weight':'bold','color':'red','size':17}

fig = plt.figure(figsize=(14,14))
plt.scatter(readworkbook['MES/AÑO'],readworkbook[2018])
plt.scatter(readworkbook['MES/AÑO'],readworkbook[2019])
plt.scatter(readworkbook['MES/AÑO'],readworkbook[2020])
plt.scatter(readworkbook['MES/AÑO'],readworkbook[2021])
plt.scatter(readworkbook['MES/AÑO'],readworkbook[2022])
plt.plot(readworkbook['MES/AÑO'],readworkbook[2018], label="2018")
plt.plot(readworkbook['MES/AÑO'],readworkbook[2019], label="2019")
plt.plot(readworkbook['MES/AÑO'],readworkbook[2020], label="2020")
plt.plot(readworkbook['MES/AÑO'],readworkbook[2021], label="2021")
plt.plot(readworkbook['MES/AÑO'],readworkbook[2022], label="2022")
plt.legend(loc="upper left")
plt.xlabel('MES', fontdict=fuente1) 
plt.ylabel('PRECIO Q', fontdict=fuente1)
plt.title("Gasolina Superior")
plt.grid
plt.show()

#regresion lineal

#Leer excel
datos = pd.read_excel("Regresion.xlsx")
datos.head()

#Convierte los datos en X - y
X = datos.iloc[:,0].values
y = datos.iloc[:,20].values

# convertir el array
X = X.reshape(-1,1)

#Calcular valores
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2)

lr = linear_model.LinearRegression()

lr.fit(X_train, y_train)

Y_pred = lr.predict(X_test)


#mostrar los datos por grafica
plt.scatter(X_test, y_test)
plt.plot(X_test, Y_pred, color='red', linewidth=3)
plt.xlabel('Mes')
plt.ylabel('Precio')
plt.title("Regresion Lineal")
plt.show()

print('Ecuacion del modelo:')
print('y = ',lr.coef_, ' x = ', lr.intercept_)

print('Precision del modelo:')
print(lr.score(X_train, y_train))